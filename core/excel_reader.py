"""
Excel file reading module.
Supports reading .xlsx files using openpyxl and Polars.
Handles large files efficiently with lazy loading.
"""

from pathlib import Path
from typing import Optional
import polars as pl
import re
from openpyxl import load_workbook
from loguru import logger


class ExcelReader:
    """Reads Excel files using Polars for performance."""

    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self.dataframe: Optional[pl.DataFrame] = None
        self.sheet_names: list = []

    def get_sheet_names(self) -> list:
        """Get all available sheet names in the Excel file."""
        try:
            wb = load_workbook(self.filepath, read_only=True, data_only=True)
            self.sheet_names = wb.sheetnames
            wb.close()
            logger.info(f"Found {len(self.sheet_names)} sheets in {self.filepath.name}")
            return self.sheet_names
        except Exception as e:
            logger.error(f"Failed to read sheet names: {e}")
            return []

    def read_sheet(self, sheet_name: str = None) -> pl.DataFrame:
        """
        Read a sheet into a Polars DataFrame.
        Uses lazy evaluation for better memory efficiency with large files.
        """
        try:
            if sheet_name is None:
                sheet_name = self.sheet_names[0] if self.sheet_names else "Sheet1"

            logger.info(f"Reading sheet '{sheet_name}' from {self.filepath.name}")

            # Read with Polars (automatically handles .xlsx)
            df = pl.read_excel(
                self.filepath,
                sheet_name=sheet_name,
            )

            # Sanitize all string columns to remove carriage returns and trailing spaces
            # This prevents _x000D_ artifacts in display and export
            df = self._sanitize_dataframe(df)

            # Add a stable per-row hash column for fast mapping of edits/filters
            # Use Polars hash_rows (deterministic) and store as string to match UI keys
            try:
                row_hash_series = df.hash_rows(seed=0).cast(pl.Utf8)
                df = df.with_columns(row_hash_series.alias("_row_hash"))
            except Exception as e:
                # Fallback: build a concatenated string and hash via Polars
                try:
                    concat = pl.concat_str([pl.col(c).cast(pl.Utf8).fill_null("") for c in df.columns], separator="|")
                    df = df.with_columns(pl.when(concat.is_not_null())
                                          .then(concat)
                                          .otherwise("")
                                          .alias("__concat_tmp__"))
                    # Hash via Python apply as last resort (may be slower on huge data)
                    df = df.with_columns(
                        pl.col("__concat_tmp__").map_elements(lambda s: str(hash(s))).alias("_row_hash")
                    ).drop("__concat_tmp__")
                except Exception as e2:
                    logger.warning(f"Failed to create _row_hash column: {e} / {e2}")

            self.dataframe = df
            logger.info(f"Loaded {len(df)} rows, {len(df.columns)} columns")
            return df

        except Exception as e:
            logger.error(f"Failed to read Excel file: {e}")
            raise

    def get_column_info(self) -> dict:
        """Get metadata about columns (names, data types)."""
        if self.dataframe is None:
            return {}

        return {
            col: str(dtype)
            for col, dtype in zip(self.dataframe.columns, self.dataframe.schema.values())
        }

    def get_preview(self, n_rows: int = 100) -> pl.DataFrame:
        """Get first n rows as preview."""
        if self.dataframe is None:
            raise ValueError("No dataframe loaded. Call read_sheet() first.")
        return self.dataframe.head(n_rows)

    def get_statistics(self) -> dict:
        """Get basic statistics about the dataset."""
        if self.dataframe is None:
            return {}

        return {
            "total_rows": len(self.dataframe),
            "total_columns": len(self.dataframe.columns),
            "column_names": self.dataframe.columns,
            "memory_usage_mb": self.dataframe.estimated_size("mb"),
        }

    def _sanitize_dataframe(self, df: pl.DataFrame) -> pl.DataFrame:
        """Sanitize all string columns in the dataframe.
        
        Removes carriage returns (\r), control characters, and trailing spaces
        that can cause display artifacts like _x000D_ especially with Telugu text.
        """
        try:
            # Process each column
            for col in df.columns:
                # Only process string/text columns
                if df[col].dtype == pl.Utf8:
                    df = df.with_columns(
                        pl.col(col)
                        .str.replace_all('_x000D_', '')  # Remove literal _x000D_ strings first
                        .str.replace_all('\r\n', '\n')  # Replace CRLF with LF
                        .str.replace_all('\r', '')  # Remove remaining carriage returns
                        .str.strip_chars_end()  # Remove all trailing whitespace
                    )
            return df
        except Exception as e:
            logger.warning(f"Failed to sanitize dataframe: {e}")
            return df  # Return original if sanitization fails
